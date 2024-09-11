import textwrap
from openpyxl.utils import get_column_letter
from utils.system_messages import SYSTEM_MESSAGE_COMPARISON, SYSTEM_MESSAGE_COMPARISON_CHUNK
import config
from utils.openAI import OpenAIClient
from utils.send_email import EmailClient
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from utils.excel_generator import ExcelReportGenerator

class SummaryComparator:
    def __init__(self, engine="gpt-4o"):
        self.engine = engine
        self.openai_client = OpenAIClient(self.engine)
        self.email_client = EmailClient()
        self.excel_report_generator = ExcelReportGenerator()  # Excel işlemleri için yeni sınıf

    def compare_with_multiple_neighbors(self, original_file_name, original_summary, neighbors, metadata=None):
        combined_neighbors_summary = "\n\n".join([neighbor['summary'] for neighbor in neighbors])
        neighbor_urls = [neighbor.get('url', '#') for neighbor in neighbors]

        combined_comparison = self.compare_summaries(
            original_summary=original_summary,
            neighbor_summary=combined_neighbors_summary,
            system_messages=SYSTEM_MESSAGE_COMPARISON,
            accumulate=True
        )

        individual_comparisons = []
        for neighbor in neighbors:
            comparison_result = self.compare_summaries(
                original_summary=original_summary,
                neighbor_summary=neighbor['summary'],
                system_messages=SYSTEM_MESSAGE_COMPARISON_CHUNK,
                accumulate=True
            )
            individual_comparisons.append(comparison_result)

        metadata_dict = {
            'combined_comparison': combined_comparison,
            'individual_comparisons': individual_comparisons,
            'keyword': metadata.get('keyword', 'N/A') if metadata else 'N/A',
            'url': metadata.get('URL', 'N/A') if metadata else 'N/A',
            'date': metadata.get('notified_date', 'N/A') if metadata else 'N/A',
            'neighbor_urls': neighbor_urls
        }

        # Use the new ExcelReportGenerator class to create the Excel report
        excel_file_path = 'comparison_report.xlsx'
        self.excel_report_generator.create_excel(metadata_dict, excel_file_path)

        # Send the email with the report
        subject = f"Summary Comparison Results for {original_file_name} vs Neighbors"
        body = "Please find attached the comparison report in Excel format."
        self.email_client.send_email(subject, body, 'recipient@example.com', excel_file_path)

    def compare_summaries(self, original_summary, neighbor_summary, system_messages, accumulate=False):
        """
        Compares the original summary with a single neighbor summary using OpenAI in a single comparison.
        Includes the name of the neighbor PDF in the comparison for better context.

        :param original_file_name: The file name for the original PDF.
        :param original_summary: A dictionary containing 'summary' and 'pdf_name' for the original PDF.
        :param neighbor_file_name: The file name for the neighbor PDF.
        :param neighbor_summary: A dictionary containing 'summary' and 'pdf_name' for the nearest neighbor PDF.
        :param accumulate: Boolean flag to determine whether to return sections for aggregation.
        :return: The combined differences between the original summary and the neighbor summary, including PDF names.
        """

        input_text = (
            f"Original Summary:\n{original_summary}\n\n"
            f"Neighbor Summary:\n{neighbor_summary}\n\n"
            f"Please provide the key differences between the original summary and the neighbor summary."
        )

        comparison_result = self.openai_client.compare_texts(input_text, system_messages)

        # sections = self.parse_comparison_result(comparison_result)

        if accumulate:
            return comparison_result

    def parse_comparison_result(self, comparison_result):
        """
        Parses the comparison result and extracts only the key differences between the original and neighbor summaries.

        :param comparison_result: The raw comparison result from OpenAI.
        :return: A list of key differences between the original and neighbor summaries.
        """
        differences = []

        # Iterate through the comparison result line by line
        for row in comparison_result.split("\n"):
            row = row.strip()

            # Capture only the lines that mention differences or changes
            if row.startswith("-"):
                differences.append(row)

        # Return the differences as a list in HTML format
        return "<ul>" + "".join(f"<li>{difference}</li>" for difference in differences) + "</ul>"

    def wrap_text(self, text, width=100):
        """
        Wraps text after a certain width without splitting words.

        :param text: The text to wrap.
        :param width: The maximum number of characters per line.
        :return: Wrapped text with new lines, preserving whole words.
        """
        return textwrap.fill(text, width=width)  # Use textwrap.fill to wrap text without splitting words

    def create_excel(self, metadata, file_name='comparison_report.xlsx'):
        """
        Creates an Excel file with the provided metadata, including clickable links for the Original Document and Neighbor PDFs,
        ensures the correct alignment of column headers, applies borders, and adjusts the column width dynamically for Key Differences.

        :param metadata: The metadata dictionary containing comparison information.
        :param file_name: The name of the output Excel file.
        :return: None
        """

        # Gri ve beyaz renkleri tanımlayın (şeffafımsı renkler)
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Daha açık gri renk
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Beyaz renk

        # Verileri hazırlayın
        data = {
            'İlgili Direktörlük': ['Çevre'],  # İlgili Direktörlük
            'Keyword': [metadata.get('keyword', 'N/A')],  # İlgili Keyword
            'Date': [metadata.get('date', 'N/A')],  # Güncelleme Tarihi
            'Kaynak': [''],  # Kaynak başlığına Original Document linki ekleyeceğiz
            'Key Differences': [self.wrap_text(metadata.get('combined_comparison', 'N/A'))]  # Key Differences
        }

        # DataFrame'e dönüştürme
        df = pd.DataFrame(data)

        # Komşu PDF linklerini yan yana sütun olarak eklemek için
        neighbor_urls = metadata.get('neighbor_urls', [])
        individual_comparisons = metadata.get('individual_comparisons', [])

        for idx, (neighbor_url, comparison_text) in enumerate(zip(neighbor_urls, individual_comparisons), 1):
            comparison_column = f'Benzer Doküman {idx}'
            df[comparison_column] = [f"{self.wrap_text(comparison_text)}\n\nLink"]

        # ExcelWriter ile Excel dosyasını oluşturun
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, startrow=0)  # startrow=0 olarak ayarlandı

            # Excel çalışma kitabını ve sayfasını alın
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Başlık hücrelerini koyu mavi yapmak için fill stili oluşturun (daha koyu mavi: 1E90FF)
            header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")

            # **Font Customization**
            header_font = Font(bold=True, color="FFFFFF", size=12, name='Arial')  # Changed the font to Arial

            # Ince kenarlık
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            # Başlıkları boyayın, hizalayın ve kenarlık ekleyin
            for col in range(1, len(df.columns) + 1):  # İlk başlıklar için sütunlar
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill  # Koyu mavi arka plan rengi
                cell.alignment = Alignment(horizontal='center', vertical='center')  # Başlıkları yatay ve dikey ortala
                cell.font = header_font  # Başlık fontu
                cell.border = thin_border  # Başlıklara kenarlık ekle

            # "Kaynak" başlığını doğru şekilde yerleştirin
            worksheet.cell(row=1, column=4).value = "Kaynak"  # 4. sütun 'Kaynak' başlığı

            # Kaynak sütunundaki Original Document linkini tıklanabilir hale getirin
            link_cell = worksheet.cell(row=2, column=4)  # 'Kaynak' sütunu 4. sütun
            original_url = metadata.get('url', '#')
            link_cell.value = 'Original Document'
            link_cell.hyperlink = original_url  # Hiperlink olarak ekleyin

            # **Font Customization for Links**
            link_font = Font(color="0000FF", underline="single", name='Calibri',
                             size=11)  # Changed link font to Calibri
            link_cell.font = link_font  # Apply the custom link font
            link_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            link_cell.border = thin_border  # Link hücresine kenarlık ekleyin

            # Alternatif renklerle sütunları boyayın (Başlıklar sabit, sütunlar bir gri bir beyaz olacak)
            for col_idx in range(1, 6):  # Tüm sütunlar için işlem yapıyoruz
                # Sütun renklendirmesi (gri ve beyaz dönüşümlü)
                if col_idx % 2 == 0:
                    fill_color = white_fill  # Çift sütunlar gri
                else:
                    fill_color = gray_fill# Tek sütunlar beyaz

                # Sütun boyunca her hücreye aynı renklendirme uygulanacak
                for row_idx in range(2,
                                     worksheet.max_row + 1):  # 2. satırdan veri satırlarına kadar olan kısmı ele alıyoruz
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = fill_color  # Alternatif renklendirme
                    cell.border = thin_border  # Kenarlık ekle
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Komşu PDF linklerini ve karşılaştırma metinlerini ekleyin
            for idx, (neighbor_url, comparison_text) in enumerate(zip(neighbor_urls, individual_comparisons), 1):
                # Hücreleri birleştirin ve başlığı ekleyin
                worksheet.merge_cells(start_row=1, start_column=5 + 2 * idx - 1, end_row=1, end_column=5 + 2 * idx)
                merged_header_cell = worksheet.cell(row=1, column=5 + 2 * idx - 1)
                merged_header_cell.value = f"Benzer Doküman {idx}"
                merged_header_cell.fill = header_fill  # Başlık hücre stili
                merged_header_cell.font = header_font  # Başlık fontu
                merged_header_cell.alignment = Alignment(horizontal='center', vertical='center')
                merged_header_cell.border = thin_border

                # Satırların alternatif olarak renklendirilmesi (gri ve beyaz)
                if idx % 2 == 0:
                    fill_color = gray_fill  # Çift satırlar gri
                else:
                    fill_color = white_fill  # Tek satırlar beyaz

                # Link hücresine hiperlink ekleyin (ilk sütunda)
                link_cell = worksheet.cell(row=2, column=5 + 2 * idx - 1)
                link_cell.value = 'Link'
                link_cell.hyperlink = neighbor_url  # Hiperlink olarak URL'yi ekleyin
                link_cell.font = link_font  # Hiperlink fontunu ekleyin
                link_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                link_cell.border = thin_border  # Link hücresine kenarlık ekleyin
                link_cell.fill = fill_color

                # Karşılaştırma metnini ilgili hücreye ekleyin (hyperlink olmadan)
                comparison_cell = worksheet.cell(row=2, column=5 + 2 * idx)
                comparison_cell.value = self.wrap_text(comparison_text)  # Sadece düz metin
                comparison_cell.font = Font(name='Arial', size=11)  # Normal font
                comparison_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                comparison_cell.border = thin_border  # Kenarlık ekleyin
                comparison_cell.fill = fill_color

                # Genişlikleri sabitleyin (çok uzun olmasını engellemek için)
                worksheet.column_dimensions[get_column_letter(5 + 2 * idx - 1)].width = 15  # Link sütunu genişliği
                worksheet.column_dimensions[get_column_letter(5 + 2 * idx)].width = 50  # Karşılaştırma sütunu genişliği

            # İçerikleri yatay ve dikey olarak ortalamak için tüm hücrelere hizalama ve kenarlık ekleyin
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1,
                                           max_col=worksheet.max_column):
                for cell in row:
                    # Eğer sütun 'E' veya 'E'den önceki sütunsa, yatay ve dikey ortalama işlemini uygula
                    if cell.column_letter < 'E':
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    else:
                        # 'E' sütunundan sonraki sütunlarda yalnızca dikey ortalama yap
                        cell.alignment = Alignment(vertical='center', wrap_text=True)
                    cell.border = thin_border  # Tüm hücrelere kenarlık ekle

            # E F G H I J K L sütunları için satır içindeki satır uzunluklarını kontrol edip sütun genişliğini ayarlayın
            for idx in range(5, 6):  # E sütunundan L sütununa kadar olan sütunlar (5. sütun = 'E')
                col_letter = get_column_letter(idx)
                max_length = 0

                for cell in worksheet[col_letter]:
                    if cell.value:
                        # Hücredeki değeri wrap_text fonksiyonu ile işleyin
                        wrapped_text = self.wrap_text(str(cell.value))

                        # İşlenmiş metindeki her satırın uzunluğunu bul ve en uzun satırı seç
                        max_length = max(max_length, max(len(line) for line in wrapped_text.split('\n')))

                # Sütun genişliğini en uzun satıra göre ayarla
                worksheet.column_dimensions[col_letter].width = max_length

            # Diğer sütun genişliklerini ayarlayın (E'den önceki sütunlar için)
            for idx, col in enumerate(worksheet.columns, 1):
                max_length = 0
                col_letter = get_column_letter(idx)
                if col_letter < 'E':  # Sadece A, B, C, D sütunları için geçerli olacak
                    for cell in worksheet[col_letter]:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    worksheet.column_dimensions[col_letter].width = max_length + 8

