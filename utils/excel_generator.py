import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import textwrap


class ExcelReportGenerator:
    """
    A class to handle the creation of Excel reports with comparison data, including hyperlinks, alternating row colors,
    and specific column formatting for the 'E' column and others.
    """

    def __init__(self):
        # Define gray and white fills for alternating row colors
        self.gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        self.header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12, name='Arial')
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                  bottom=Side(style='thin'))
        self.link_font = Font(color="0000FF", underline="single", name='Calibri', size=11)
        self.custom_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    def apply_header_style(self, worksheet, col_count):
        # Başlıkları biçimlendiren fonksiyon
        for col in range(1, col_count + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = self.header_font
            cell.border = self.thin_border

    def format_data_columns(self, worksheet, max_row):
        # Alternatif hücre renklendirmesi ve kenarlık işlemleri
        for col_idx in range(1, 6):
            fill_color = self.white_fill if col_idx % 2 == 0 else self.gray_fill
            for row_idx in range(2, max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill = fill_color
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def adjust_first_four_column_width(self, worksheet):
        # İlk 4 sütunun genişliğini sabit olarak ayarlayın
        worksheet.column_dimensions['A'].width = 20  # Direktörlük
        worksheet.column_dimensions['B'].width = 20  # Keyword
        worksheet.column_dimensions['C'].width = 20  # Date
        worksheet.column_dimensions['D'].width = 20  # Kaynak (Original Document linki)
        worksheet.column_dimensions['E'].width = 20  # Kaynak (Original Document linki)

    def add_combined_neighbor_links_and_comparisons_with_comments(self, worksheet, neighbor_urls,
                                                                  individual_comparisons):
        """
        Tek bir hücreye tüm benzer dokümanların linklerini ve karşılaştırma metinlerini ekleyen fonksiyon.
        Hücreye farklı bir arka plan rengi uygulanır.
        """
        # Benzer Dokümanlar hücresini ekleyin
        combined_cell = worksheet.cell(row=2, column=6)  # Tek bir hücrede olacak (örneğin F2)
        combined_cell.value = "Benzer Dokümanlar"
        combined_cell.font = self.link_font
        combined_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        combined_cell.border = self.thin_border
        combined_cell.fill = self.custom_fill  # Hücreye arka plan rengi ekleme

        # Tüm benzer dokümanların detaylarını hazırlayın, başlık ekleyin
        combined_comment_text = "Benzer Dokümanlar:\n\n"  # Yorumun başlığı
        for idx, (neighbor_url, comparison_text) in enumerate(zip(neighbor_urls, individual_comparisons), 1):
            # Her bir dokümanın detaylarını alt alta listeleyin (linkler sarılmaz)
            combined_comment_text += f"\nBenzer Doküman {idx}: {neighbor_url}\nKarşılaştırma: {comparison_text}\n\n"

        # Hücreye hazırlanan metni yorum olarak ekleyin
        combined_comment = Comment(combined_comment_text, "Comparison")
        combined_cell.comment = combined_comment

        # Sütun genişliğini ayarlama
        worksheet.column_dimensions[get_column_letter(6)].width = 30  # Benzer Dokümanlar sütunu genişliği

    def create_excel(self, metadata, file_name='comparison_report.xlsx'):
        # Verileri hazırlayın
        data = {
            'İlgili Direktörlük': ['Çevre'],
            'Keyword': [metadata.get('keyword', 'N/A')],
            'Date': [metadata.get('date', 'N/A')],
            'Kaynak': [''],
            'Key Differences': ['...']
        }

        df = pd.DataFrame(data)
        neighbor_urls = metadata.get('neighbor_urls', [])
        individual_comparisons = metadata.get('individual_comparisons', [])

        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, startrow=0)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Başlık biçimlendirme
            self.apply_header_style(worksheet, len(df.columns))

            # "Kaynak" başlığını ekleyin ve Original Document linkini tıklanabilir hale getirin
            worksheet.cell(row=1, column=4).value = "Kaynak"
            link_cell = worksheet.cell(row=2, column=4)
            original_url = metadata.get('url', '#')
            link_cell.value = 'Original Document'
            link_cell.hyperlink = original_url
            link_cell.font = self.link_font
            link_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            link_cell.border = self.thin_border

            # "Key Differences" için yorum ekleyin
            key_diff_cell = worksheet.cell(row=2, column=5)  # "Key Differences" hücresi
            key_diff_cell.value = "..."
            key_diff_cell.font = Font(name='Arial', size=11)
            key_diff_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            key_diff_cell.border = self.thin_border

            # "Key Differences" yorum ekleme
            key_diff_comment = Comment(metadata.get('combined_comparison', 'N/A'), "Key Differences")
            key_diff_cell.comment = key_diff_comment

            # Veri sütunlarını biçimlendirme (alternatif renklendirme ve kenarlık)
            self.format_data_columns(worksheet, worksheet.max_row)

            # Komşu linkleri ve karşılaştırma metinlerini ekleyin, metinleri yorum olarak ekleyin
            self.add_combined_neighbor_links_and_comparisons_with_comments(worksheet, neighbor_urls,
                                                                           individual_comparisons)

            # İlk 4 sütunun genişliğini ayarla
            self.adjust_first_four_column_width(worksheet)

